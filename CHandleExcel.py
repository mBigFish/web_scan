import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class HandleExcel:
    def __init__(self, file_path):
        self.file_path = file_path

    def find_empty_start_column(self):
        try:
            df = pd.read_excel(self.file_path)
            num_columns = len(df.columns)
            for i in range(num_columns):
                column_data = df.iloc[:, i]
                if column_data.isnull().all() or all(str(value).strip() == '' for value in column_data):
                    return True, i + 2
            return True, num_columns + 2
        except Exception as e:
            print(f'读取Excel文件出错：{str(e)}')
            return False, f'读取Excel文件出错：{str(e)}'

    def read_urls_from_excel(self, column_name=False):
        try:
            df = pd.read_excel(self.file_path)
            url_column = None
            if column_name:
                for col in df.columns:
                    if col.lower() == column_name:
                        url_column = col
                        break
            else:
                for col in df.columns:
                    if col.lower() == "url":
                        url_column = col
                        break
                    elif col.lower() == 'link':
                        url_column = col
                        break
            if url_column is None:
                print("未找到包含关键字的列的列")
                return False, "未找到包含关键字的列的列"

            urls_dict = {}
            for index, value in df[url_column].items():
                urls_dict[index + 2] = value  # 行号从0开始，所以要加上2
                # print(value)

            return True, urls_dict

        except Exception as e:
            print(f'读取Excel文件出错：{str(e)}')
            return False, f'读取Excel文件出错：{str(e)}'

    def write_excel_header(self, header_list, start_col):
        try:
            wb = load_workbook(self.file_path)
            sheet = wb.active
            for idx, header in enumerate(header_list):
                col_letter = get_column_letter(start_col + idx)
                sheet[f"{col_letter}1"] = header

            for col in range(1, start_col + len(header_list)):
                col_letter = get_column_letter(col)
                cell = sheet[f"{col_letter}1"]
                cell.font = Font(bold=True)

            wb.save(self.file_path)
            return True
        except Exception as e:
            print(f"写入Excel文件出错：{e}")
            return False

    def write_excel_data(self, data_list, start_col):
        try:
            wb = load_workbook(self.file_path)
            sheet = wb.active

            for row_letter in data_list:
                data_list_row = data_list[row_letter]
                for idx, data_li in enumerate(data_list_row):
                    col_letter = get_column_letter(start_col + idx)
                    sheet[f"{col_letter}{row_letter}"] = data_li

            wb.save(self.file_path)
            return True
        except Exception as e:
            print(f"写入Excel文件出错：{e}")
            return False
