# -*- coding: utf-8 -*-
import json
import os
import random
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor
from fake_useragent import UserAgent
import asyncio
import httpx
import openpyxl
import requests
import tldextract
from PySide6.QtCore import QThread, Signal, QUrl
from PySide6.QtGui import QDesktopServices
from bs4 import BeautifulSoup
import re
from PySide6.QtWidgets import QApplication, QWidget, QFileDialog, QMessageBox, QButtonGroup, QRadioButton
from openpyxl.styles import Font
from ui import Ui_Form
from qt_material import apply_stylesheet
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ----------------------------------------全部变量定义块--------------------------------------------------------------------------

"""UA"""
# headers = {
#     'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36'
# }

"""网站错误关键词"""
Errors = ['tengine', 'Apache Tomcat', '站点创建成功', '不存在', '访问报错', 'Domain has expired', '网站建设中',
          '官网登录入口', '502', '网站维护', '温馨提示', '无标题文档', '阻断页面', 'CentOS', '阻止', '无法访问', '域名',
          '站点已暂停', '404', '没有找到站点', '未获取到网站标题', '到期', 'nginx', 'IIS', ]

"""网站违法关键词"""
Illegal = ['综合体育', '安全加密检测', '安全检测..', '无码', 'A片', '官方入口', '在线体育', '半岛', '体育彩票',
           '太阳成集团',
           'ios/安卓/手机版app', '官网(中国)', '快三官网', '金博体育', '(中国)官方网站', '真人下注', 'Loading....',
           '体育(中国)',
           'ios', '官网登录入口', 'bwin必赢', '太阳商城', '中欧体育', '愉拍', '日本', '澳门', 'OB体育', '开云',
           'Im体育',
           '必威betway', '亚博', 'AV', '彩票',',好吊视频', '一区二区三区', '国产SUV', '久久蜜', '精品日产', '麻豆',
           '皇冠体育', '三级黄色', '茄子视频', '视频色版', '威尼斯', '小鸡鸡', '骚逼逼', '视频污版', '欧美', '性爽',
           '硬汉视频', '性爱', '人妻', '少妇', '精品视频', '污污', '香蕉视频', '喷水', '啪啪', '91', '污视频', '荔枝视频']

"""全局变量定义"""
EXCEL_READ_PATH = ''  # excel文件路径
EXCEL_WRITE_PATH_RELATIVE = 'output.xlsx'   # Excel保存文件相对路径
EXCEL_WRITE_PATH_ABSOLUTE = ''  # Excel保存文件绝对路径
API_URL = "http://www.jucha.com/item/search"


ALL_URLS_COUNT = 0     # 所有urls数量
NOW_URLS_COUNT = 1     # 目前已爬取urls数量

SPIDER_VALUES_DIR = {}
ICP_VALUES_DIR = {}
PS_VALUES_DIR = {}


output_head_value = ["域名", "判定结果", "网站标题", "状态码", "底部悬挂icp备案", "底部悬挂网安备案"]
output_api_icp_head_value = ['icp备案当前状态', 'icp备案审核时间', '单位名称', '单位性质', '备案号', '网站名称', '网站首页地址', '网站负责人']
output_api_ps_head_value = ['网安备案当前状态', '网站类别', '开办者主体', '备案地公安机关', '联网备案时间', '公安备案号']

"""参数设置"""
# API调用功能
API_FUNCTION = False
# 线程数
p_max_workers = 2000
# 超时时间 timeout
p_timeout = 5
# 认证SSL证书开关 verify
p_verify = False
# 重定向开关 allow_redirects
p_allow_redirects = True


# ----------------------------------------窗口代码块--------------------------------------------------------------------------
class MyWindow(QWidget, Ui_Form):
    def __init__(self):
        super().__init__()
        self.api_button_group = None
        self.allow_redirect_button_group = None
        self.verify_button_group = None
        self.worker_thread = None
        self.setupUi(self)
        self.bind()

    def bind(self):

        # 创建工作线程
        self.worker_thread = WorkerThread()
        # 进度条初始值
        self.progress_bar.setValue(0)
        # 连接信号与槽
        self.worker_thread.update_signal.connect(self.update_text_edit)
        self.worker_thread.update_progress_signal.connect(self.update_progress_bar)
        self.worker_thread.worker_thread_finsh.connect(self.worker_thread_finsh)

        """主页面组件"""
        # 开始运行按钮
        self.start_btn.clicked.connect(self.start_btn_func)
        # 结束运行按钮
        self.exit_btn.clicked.connect(self.exit_btn_func)
        # 选择表格 读取urls
        self.excel_select_btn.clicked.connect(self.excel_select_btn_func)
        # 运行窗口提示
        self.operation_status_edit.setPlaceholderText("请选择Excel表格时，URL一列放在表格第一列，且从第一行开始，当程序运行完成时，将会为您输出保存文件路径！\n"
                                                      "如需具体教程，请在更多关于中查看使用介绍")
        # 打开文件按钮
        self.open_file_btn.setEnabled(False)
        self.open_file_btn.setText("未运行")
        self.open_file_btn.clicked.connect(self.open_file_btn_func)
        # 超时时间
        self.timeout_slider.setValue(p_timeout)
        self.timeout_lab.setText(str(p_timeout))
        self.timeout_slider.valueChanged.connect(self.timeout_slider_value_changed)
        # 最大线程数
        self.max_workers_slider.setValue(p_max_workers)
        self.max_workers_lab.setText(str(p_max_workers))
        self.max_workers_slider.valueChanged.connect(self.max_workers_slider_value_changed)
        # 认证SSL证书开关
        self.verify_button_group = QButtonGroup()
        self.verify_button_group.addButton(self.verify_yes_radio_btn)
        self.verify_button_group.addButton(self.verify_no_radio_btn)
        self.verify_yes_radio_btn.setChecked(True) if p_verify else self.verify_no_radio_btn.setChecked(True)
        self.verify_button_group.buttonClicked.connect(self.verify_group_value_changed)
        # 重定向开关
        self.allow_redirect_button_group = QButtonGroup()
        self.allow_redirect_button_group.addButton(self.allow_redirect_yes_btn)
        self.allow_redirect_button_group.addButton(self.allow_redirect_no_btn)
        self.allow_redirect_yes_btn.setChecked(True) if p_allow_redirects else self.verify_no_radio_btn.setChecked(True)
        self.allow_redirect_button_group.buttonClicked.connect(self.allow_redirect_button_group_value_changed)
        # API调用功能开关
        self.api_button_group = QButtonGroup()
        self.api_button_group.addButton(self.api_yes_radio_btn)
        self.api_button_group.addButton(self.api_no_radio_btn)
        self.api_yes_radio_btn.setChecked(True) if API_FUNCTION else self.api_no_radio_btn.setChecked(True)
        self.api_button_group.buttonClicked.connect(self.api_button_group_value_changed)

        """第三页组件"""
        # 检测更新按钮
        self.update_btn.clicked.connect(self.check_for_updates)
        # 软件介绍按钮
        self.introduction_use_btn.clicked.connect(lambda:self.open_url_in_browser("https://blog.mbigfish.com/index.php/archives/3064/"))
        # 开源地址按钮
        self.open_source_address_btn.clicked.connect(lambda:self.open_url_in_browser("https://github.com/mBigFish/web_scan"))

    """API功能调用选项框函数"""
    def api_button_group_value_changed(self):
        selected_button = self.api_button_group.checkedButton()
        global API_FUNCTION
        API_FUNCTION = True if selected_button.text() == "开启" else False

    """请求响应， timeout_slider函数"""
    def timeout_slider_value_changed(self, value):
        global p_timeout
        p_timeout = value
        self.timeout_lab.setText(str(value))

    """最大线程数， max_workers__slider函数"""
    def max_workers_slider_value_changed(self, value):
        global p_max_workers
        p_max_workers = value
        self.max_workers_lab.setText(str(value))

    """重定向选项框函数"""
    def allow_redirect_button_group_value_changed(self):
        selected_button = self.allow_redirect_button_group.checkedButton()
        global p_allow_redirects
        p_allow_redirects = True if selected_button.text() == "开启" else False

    """认证SSL证书选项框函数"""
    def verify_group_value_changed(self):
        selected_button = self.verify_button_group.checkedButton()
        global p_verify
        p_verify = True if selected_button.text() == "开启" else False

    """启动工作线程函数"""
    def start_worker_thread(self):
        # 启动工作线程
        self.worker_thread.start()

    """打开文件按钮"""
    @staticmethod
    def open_file_btn_func(self):
        # mac打开文件
        # subprocess.run(["open", EXCEL_WRITE_PATH_ABSOLUTE])
        # windows打开文件
        subprocess.run(["start", EXCEL_WRITE_PATH_ABSOLUTE], shell=True)

    """工作线程完成时函数"""
    def worker_thread_finsh(self):
        self.start_btn.setText("运行完成")
        self.exit_btn.setText("退出程序")
        self.exit_btn.setEnabled(True)
        self.open_file_btn.setText("打开文件")
        self.open_file_btn.setEnabled(True)

    """"实时更新文本框显示内容函数"""
    def update_text_edit(self, message):
        self.operation_status_edit.append(message)

    """更新进度条函数"""
    def update_progress_bar(self, value):
        self.progress_bar.setValue(value)

    """选择表格按钮函数"""
    def excel_select_btn_func(self):
        # 打开文件对话框以获取Excel文件路径
        file_dialog = QFileDialog()
        file_path = file_dialog.getOpenFileName(self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls);;所有文件 (*)")
        self.excel_input_edit.setText(file_path[0])
        global EXCEL_READ_PATH
        EXCEL_READ_PATH = file_path[0]

    """开始按钮函数"""
    def start_btn_func(self):
        if not EXCEL_READ_PATH:
            QMessageBox.information(self, "提示", "请选择表格！")
        else:
            self.check_path_exists()
            self.start_worker_thread()
            self.start_btn.setEnabled(False)
            self.start_btn.setText("正在运行")
            self.exit_btn.setEnabled(False)
            self.exit_btn.setText("切勿退出")
            self.open_file_btn.setText("请稍后")

    """结束运行按钮函数"""
    @staticmethod
    def exit_btn_func(self):
        QApplication.instance().quit()

    """检测文件是否存在函数"""
    def check_path_exists(self):
        warning_message = "检测到当前存在output.xlsx输出文件，请将该文件改名或移动到其它文件后运行程序，若执意运行，会导致该文件被新数据覆盖，点击yes即可执意运行"
        if os.path.exists(EXCEL_WRITE_PATH_ABSOLUTE):
            reply = QMessageBox.question(self, '警告', warning_message, QMessageBox.StandardButton.Yes,
                                         QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                pass
            else:
                self.exit_btn_func(self)

    """检测更新函数"""
    def check_for_updates(self):
        # 向服务器请求最新版本信息
        try:
            response = requests.get('https://m.mbigfish.com/main/WebCheckUpdateDetection.php', params={'version': '1.0.2'})
            response.raise_for_status()  # 检查请求是否成功
            latest_version_info = response.json()

            # 检查是否有新版本
            if self.isNewVersionAvailable(latest_version_info):
                # 有新版本可用
                update_message = (f"发现新版本 {latest_version_info['version']}\n"
                                  f"更新时间为{latest_version_info['release_date']}\n 是否立即更新？")

                reply = QMessageBox.question(self, '更新提示', update_message, QMessageBox.StandardButton.Yes, QMessageBox.StandardButton.No)
                if reply == QMessageBox.StandardButton.Yes:
                    url_to_open = latest_version_info['download_url']
                    self.open_url_in_browser(url_to_open)
            else:
                # 已经是最新版本
                QMessageBox.information(self, "更新检查", f"当前已经是最新版本{latest_version_info['version']}\n "
                                                         f"更新时间为{latest_version_info['release_date']}\n"
                                                         f"暂时无需更新")
        except requests.RequestException as e:
            # 处理请求异常
            QMessageBox.warning(self, "更新检查失败", f"检查更新时发生错误\n 请前往GitHub地址查看版本更新情况！")

    """通过使用默认浏览器打开链接"""
    @staticmethod
    def open_url_in_browser(url_to_open):
        QDesktopServices.openUrl(QUrl(url_to_open))

    """比较版本函数"""
    @staticmethod
    def isNewVersionAvailable(latest_version_info):
        # 比较当前版本与最新版本
        current_version = "1.0.2"  # 你的应用当前版本号
        return latest_version_info["version"] > current_version


# ----------------------------------------工作线程代码块--------------------------------------------------------------------------
class WorkerThread(QThread):
    update_signal = Signal(str)
    update_progress_signal = Signal(int)
    worker_thread_finsh = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.urls = []

    def run(self):
        # 批量检测前的显示
        self.update_signal.emit("-"*42 + "程序正在开始运行" + "-"*42 + '\n')
        # 从公表格中读取所有URL
        self.urls = read_excel()
        start_time = time.time()
        # 运行主函数
        self.main()
        finsh_time = time.time()
        # 批量检测终止时的显示
        self.update_signal.emit("-"*42 + "程序已经运行完成" + "-"*42 + '\n')
        self.update_signal.emit("-" * 44 + "重要信息输出" + "-" * 44 + '\n')
        self.update_signal.emit(f"程序耗时：{finsh_time - start_time}")
        self.update_signal.emit(f"网站总数：{len(self.urls)}")
        self.update_signal.emit(f"爬取次数：{ALL_URLS_COUNT}")
        self.update_signal.emit(f"保存路径：{EXCEL_WRITE_PATH_ABSOLUTE}")
        self.update_signal.emit("\n" + "-" * 44 + "重要信息输出" + "-" * 44 + '')

    """API异步主函数"""
    @staticmethod
    async def api_main_async(api_tasks):
        api_task_list = split_into_lists(api_tasks, 100)
        for api_task_li in api_task_list:
            await asyncio.gather(*api_task_li)
        output_dir_excel(6, ICP_VALUES_DIR)
        output_dir_excel(14, PS_VALUES_DIR)

    """异步主函数"""
    def main(self):
        global ALL_URLS_COUNT
        if API_FUNCTION:
            # 获取即将要检测的URL总数量
            ALL_URLS_COUNT = len(self.urls) * 3
            output_head(6, output_api_icp_head_value)
            output_head(14, output_api_ps_head_value)
            # 开启异步任务
            url_id = 1
            api_tasks = []
            for url in self.urls:
                url_id += 1
                url = remove_http_prefix(str(url))
                url = "http://" + url
                # 添加api请求
                api_tasks.append(self.api_fetch_data("beian", url_id, url))
                api_tasks.append(self.api_fetch_data("wangan", url_id, url))
            # 调用api请求
            self.update_signal.emit("-" * 40 + "正在运行api数据爬取" + "-" * 40 + '\n')
            asyncio.run(self.api_main_async(api_tasks))
            self.update_signal.emit("-" * 40 + "已经完成api数据爬取" + "-" * 40 + '\n')
        else:
            # 获取即将要检测的URL总数量
            ALL_URLS_COUNT = len(self.urls)
        # 开启多线程任务
        output_head(0, output_head_value)
        self.update_signal.emit("-" * 40 + "正在运行批量数据爬取" + "-" * 40 + '\n')
        with ThreadPoolExecutor(max_workers=p_max_workers) as executor:
            # 提交任务给线程池
            url_id = 1
            for url in self.urls:
                url_id += 1
                url = remove_http_prefix(str(url))
                url = "http://" + url
                executor.submit(self.get_request, self, url_id, url, )
        output_dir_excel(0, SPIDER_VALUES_DIR)
        self.update_signal.emit("-" * 40 + "已经完成批量数据爬取" + "-" * 40 + '\n')
        self.worker_thread_finsh.emit()

    """异步获取URL信息函数"""
    @staticmethod
    def get_request(self, url_id, url):
        headers = get_random_user_agent()
        try:
            response = requests.get(url, headers=headers, timeout=p_timeout, verify=p_verify, allow_redirects=p_allow_redirects)
            # response.raise_for_status()
            status_code = response.status_code
            encodings = response.apparent_encoding
            if encodings:
                response.encoding = encodings
            else:
                response.encoding = 'utf-8'
            result = response.text
        except requests.exceptions.RequestException as e:
            result = f"请求发生错误，网站可能打不开"
            status_code = None
        except Exception as e:
            result = f"未知异常"
            status_code = None
        self.parse(url_id, url, status_code, result)

    """解析爬虫数据函数"""
    def parse(self, url_id, url, status_code, html):
        if status_code:
            """获取网站信息"""
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.title.string if soup.title else "不确定，建议手动查看" if status_code == 200 else '未获取到网站标题'
            title = title.strip() if hasattr(title, 'strip') else title
            """判断网站是否异常或违法"""
            web_status = "正常网站"
            if status_code == 200 and title:
                web_status = "异常网站" if any(item in title for item in Errors) else web_status
                web_status = "违法网站" if any(item in title for item in Illegal) else web_status
                web_status = "域名出售" if url in title and "官网首页" in title else web_status
            else:
                web_status = "异常网站"
            """获取备案信息"""
            # 获取icp备案
            icp_number = parse_icp(soup)
            # 获取公安备案
            public_security_number = parse_public_security(soup)
            ret_info = f"网站名称：{title}；判定结果：{web_status}"
            values = [url, web_status, title, status_code, icp_number, public_security_number]
            SPIDER_VALUES_DIR[url_id] = values
            self.update_signal.emit(ret_info)
        else:
            web_status = "错误网站"
            ret_info = f"网站名称：无名称；判定结果：{web_status}"
            self.update_signal.emit(ret_info)
            values = [url, web_status, html]
            SPIDER_VALUES_DIR[url_id] = values
        """进度条"""
        global NOW_URLS_COUNT
        NOW_URLS_COUNT += 1
        self.update_progress_signal.emit(NOW_URLS_COUNT/ALL_URLS_COUNT*100)

    """请求api函数"""
    async def api_fetch_data(self, route, url_id, url):
        headers = get_random_user_agent()
        async with httpx.AsyncClient() as client:
            domain = remove_domain_prefix(url)
            if route == "beian":
                payload = {
                    'domain': domain,
                    'items[]': 24,
                    'type': 1,
                    'route': 'beian'
                }
                self.update_signal.emit(f"URL：{url}；icp备案获取成功")
            elif route == "wangan":
                payload = {
                    'domain': domain,
                    'items[]': 3,
                    'type': 1,
                    'route': 'wangan'
                }
                self.update_signal.emit(f"URL：{url}；网安备案获取成功")
            else:
                print("类型错误！")
            response = await client.post(url=API_URL,headers=headers, data=payload, timeout=30)
            response = response.text
            if response.startswith(u'\ufeff'):
                response = response.encode('utf8')[3:].decode('utf8')
                data = json.loads(response)
            # print(response)
            parse_api(route, url_id, data)
            # 进度条
            global NOW_URLS_COUNT
            NOW_URLS_COUNT += 1
            self.update_progress_signal.emit(NOW_URLS_COUNT / ALL_URLS_COUNT * 100)


# ----------------------------------------API代码块--------------------------------------------------------------------------


"""解析分流api返回数据函数"""
def parse_api(route, url_id, data):
    if route == "beian":
        value = parse_add_icp(data)
        ICP_VALUES_DIR[url_id] = value
    elif route == "wangan":
        value = parse_add_public_security(data)
        PS_VALUES_DIR[url_id] = value
    else:
        print("请求参数错误")


"""解析api返回icp备案数据函数"""
def parse_add_icp(icp_json):
    try:
        code_main = icp_json['code']
        if code_main == 1:
            err_code = icp_json['data']['beian']['err_code']
            if err_code == 200:
                code = icp_json['data']['beian']['data']['code']
                if code == 1:
                    current_status = icp_json['data']['beian']['data']['msg']  # 当前状态
                    if not current_status == "未备案":
                        review_time = icp_json['data']['beian']['data']['data']['sj']  # 审核时间
                        company_name = icp_json['data']['beian']['data']['data']['mc']  # 单位名称
                        company_kind = icp_json['data']['beian']['data']['data']['lx']  # 单位性质
                        icp_num = icp_json['data']['beian']['data']['data']['bah']  # 备案号
                        web_name = icp_json['data']['beian']['data']['data']['bam']  # 网站名称
                        web_home_link = icp_json['data']['beian']['data']['data']['sy']  # 网站首页地址
                        web_principal = icp_json['data']['beian']['data']['data']['fzr'] # 网站负责人
                        value = [current_status, review_time, company_name, company_kind, icp_num, web_name, web_home_link, web_principal]
                    else:
                        value = [current_status]
                else:
                    value = ["提交验证失败"]
            else:
                value = ["请求失败"]
        else:
            msg = icp_json['msg']
            value = [msg]
    except Exception as e:
        value = ["未知错误"]
    return value


"""解析api返回网安备案数据函数"""
def parse_add_public_security(ps_json):
    try:
        code_main = ps_json['code']
        if code_main == 1:
            err_code = ps_json['data']['wangan']['err_code']
            if err_code == 200:
                code = ps_json['data']['wangan']['data']['code']
                if code == 1:
                    current_status = ps_json['data']['wangan']['data']['msg']  # 当前状态
                    if not current_status == "未备案":
                        web_kind = ps_json['data']['wangan']['data']['data']['fs']  # 网站类别
                        founder_name = ps_json['data']['wangan']['data']['data']['name']  # 开办者主体  单位名称
                        place_filing = ps_json['data']['wangan']['data']['data']['dz']  # 备案地公安机关
                        online_filing_time = ps_json['data']['wangan']['data']['data']['sj']  # 联网备案时间
                        public_security_num = ps_json['data']['wangan']['data']['data']['mc']  # 公安备案号
                        value = [current_status, web_kind, founder_name, place_filing, online_filing_time, public_security_num]
                    else:
                        value = [current_status]
                else:
                    value = ["提交验证失败"]
            else:
                value = ['请求失败']
        else:
            msg = ps_json['msg']
            value = [msg]
    except Exception as e:
        value = ["未知错误"]
    return value

# ----------------------------------------爬虫代码块--------------------------------------------------------------------------


"""写入表格头函数"""
def output_head(col_index, values):
    try:
        workbook = openpyxl.load_workbook("./output.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    # 获取要写入的工作表
    sheet = workbook.active  # 这里假设你要操作的是默认的工作表
    col_index = col_index
    for value in values:
        row_index = 1
        col_index = col_index + 1
        sheet.cell(row=row_index, column=col_index, value=value)
        bold_font = Font(bold=True)
        sheet.cell(row=row_index, column=col_index).font = bold_font
        sheet.column_dimensions[sheet.cell(row=row_index, column=col_index).column_letter].width = 20
    # 保存工作簿
    workbook.save(EXCEL_WRITE_PATH_RELATIVE)


"""从Excel表中读取URL函数"""
def read_excel():
    urls = []
    try:
        # 打开 Excel 文件
        workbook = openpyxl.load_workbook(EXCEL_READ_PATH)
        # 选择第一个工作表
        sheet = workbook.active
        # 遍历行
        for row in sheet.iter_rows(values_only=True):
            url = row[0]  # 无http头
            urls.append(url)
    except FileNotFoundError:
        print("文件未找到，请检查文件路径是否正确。")
    except Exception as e:
        print(f"发生了其他错误: {e}")
    return urls


"""批量检测icp备案获取函数"""
def parse_icp(soup):
    # 找到包含 ICP 备案关键字的元素
    icp_elements = soup.find_all(string=re.compile(r'ICP备\d'))
    # 提取备案信息
    icp_number = "未找到icp备案"
    for element in icp_elements:
        # 进一步处理字符串，提取具体备案信息
        # icp_match = re.search(r'ICP备[^\d]*(\d+)[^\u4e00-\u9fa5]*([\u4e00-\u9fa5]+)', element)
        # icp_match = re.search(r'([\u4e00-\u9fa5]*ICP备[^\d]*\d+[^\u4e00-\u9fa5]*[\u4e00-\u9fa5]+-*\d*)', element)
        icp_match = re.search(r'([\u4e00-\u9fa5]?ICP备\d+[^\u4e00-\u9fa5]*[\u4e00-\u9fa5]+-*\d*)', element)
        if icp_match:
            icp_number = icp_match.group(1)
        else:
            icp_number = "可能存在icp备案，但未爬取到"
    icp_number = icp_number.replace(" ", "")
    return icp_number


"""批量检测公安备案获取函数"""
def parse_public_security(soup):
    # 找到包含公安备案关键字的元素
    public_security_elements = soup.find_all(string=re.compile(r'公网安备'))
    # 提取公安备案信息
    public_security_number = "未找到公安备案"
    for element in public_security_elements:
        # 进一步处理字符串，提取具体备案信息
        public_security_match = re.search(r'([\u4e00-\u9fa5]?公网安备\s*\d+号)', element)
        if public_security_match:
            public_security_number = public_security_match.group(1)
        else:
            public_security_number = "可能存在公安备案，但未爬取到"
    public_security_number = public_security_number.replace(" ", "")
    return public_security_number


"""列表分割函数"""
def split_into_lists(lst, chunk_size):
    """将列表按照指定大小分割成子列表"""
    for i in range(0, len(lst), chunk_size):
        yield lst[i:i + chunk_size]


"""将数据写入表格函数---废弃函数"""
def output_excel(col_index, url_id, values):
    try:
        # 尝试打开已存在的 Excel 文件
        workbook = openpyxl.load_workbook("./output.xlsx")
        sheet = workbook["Sheet"]
    except FileNotFoundError:
        # 如果文件不存在，则创建新的 Excel 工作簿
        workbook = openpyxl.Workbook()
        # 获取要写入的工作表
        sheet = workbook.active
    col_index = col_index
    for value in values:
        row_index = url_id
        col_index = col_index + 1
        sheet.cell(row=row_index, column=col_index, value=value)
    # 保存工作簿
    workbook.save(EXCEL_WRITE_PATH_RELATIVE)

def output_dir_excel(col, data_dir):
    try:
        # 尝试打开已存在的 Excel 文件
        workbook = openpyxl.load_workbook(EXCEL_WRITE_PATH_RELATIVE)
        sheet = workbook["Sheet"]
    except FileNotFoundError:
        # 如果文件不存在，则创建新的 Excel 工作簿
        workbook = openpyxl.Workbook()
        # 获取要写入的工作表
        sheet = workbook.active
    for url_id in data_dir:  # 行
        col_index = col
        for value in data_dir[url_id]:  # 列
            row_index = url_id
            col_index = col_index + 1
            sheet.cell(row=row_index, column=col_index, value=value)
    # 保存工作簿
    workbook.save(EXCEL_WRITE_PATH_RELATIVE)


"""去除http头函数"""
def remove_http_prefix(url):
    if url.startswith("http://"):
        return url[len("http://"):]
    elif url.startswith("https://"):
        return url[len("https://"):]
    else:
        return url


"""获取链接域名函数"""
def remove_domain_prefix(url):
    # 使用 tldextract 提取域名信息
    extracted_info = tldextract.extract(url)
    # 组合新的 URL，去掉子域名前缀
    new_url = "{}.{}".format(extracted_info.domain, extracted_info.suffix)
    return new_url


"""随机请求头函数"""
def get_random_user_agent():
    user_agent_list = [ \
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/22.0.1207.1 Safari/537.1" \
        "Mozilla/5.0 (X11; CrOS i686 2268.111.0) AppleWebKit/536.11 (KHTML, like Gecko) Chrome/20.0.1132.57 Safari/536.11", \
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1092.0 Safari/536.6", \
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.6 (KHTML, like Gecko) Chrome/20.0.1090.0 Safari/536.6", \
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.1 (KHTML, like Gecko) Chrome/19.77.34.5 Safari/537.1", \
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.9 Safari/536.5", \
        "Mozilla/5.0 (Windows NT 6.0) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.36 Safari/536.5", \
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 5.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3", \
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_8_0) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1063.0 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1062.0 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 6.1) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.1 Safari/536.3", \
        "Mozilla/5.0 (Windows NT 6.2) AppleWebKit/536.3 (KHTML, like Gecko) Chrome/19.0.1061.0 Safari/536.3", \
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24", \
        "Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/535.24 (KHTML, like Gecko) Chrome/19.0.1055.1 Safari/535.24"
    ]
    UserAgent=random.choice(user_agent_list)
    headers = {'User-Agent': UserAgent}
    return headers

"""主函数"""
if __name__ == '__main__':
    # 获取当前工作目录
    current_directory = os.getcwd()
    # 获取保存文件绝对路径
    EXCEL_WRITE_PATH_ABSOLUTE = os.path.join(current_directory, EXCEL_WRITE_PATH_RELATIVE)

    app = QApplication([])
    apply_stylesheet(app, theme="dark_cyan.xml")
    window = MyWindow()
    window.show()
    sys.exit(app.exec())
