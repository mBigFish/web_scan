# -*- coding: utf-8 -*-
import time
import ssl
import aiohttp
import asyncio
import openpyxl
import pandas as pd
import xlwt
from PySide6.QtCore import QThread, Signal
from bs4 import BeautifulSoup
from email.charset import Charset
import re
import urllib3
from PySide6.QtWidgets import QApplication, QMainWindow, QPushButton, QLabel, QLineEdit, QWidget, QFileDialog
from test1 import Ui_widget
import app as spider

# ----------------------------------------全部变量定义块--------------------------------------------------------------------------
"""游览器头定义"""
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
}

"""网站错误关键词"""
Errors = ['tengine', 'Apache Tomcat', '站点创建成功', '不存在', '访问报错', 'Domain has expired', '网站建设中',
          '官网登录入口', '502', '维护', '温馨提示', '无标题文档', '阻断页面', 'CentOS', '阻止', '无法访问', '域名',
          '站点已暂停', '404', '没有找到站点', '域名', 'No title found', '到期', 'nginx', 'IIS',]

"""网站违法关键词"""
Illegal = ['综合体育', '安全加密检测', '安全检测..', '无码', 'A片', '官方入口', '在线体育', '半岛', '体育彩票', '太阳成集团',
           'ios/安卓/手机版app', '官网(中国)', '快三官网', '金博体育', '(中国)官方网站', '真人下注', 'Loading....', '体育(中国)',
           'ios', '官网登录入口', 'bwin必赢', '太阳商城', '中欧体育', '愉拍', '日本', '澳门', 'OB体育', '开云', 'Im体育',
           '必威betway', '亚博', 'AV', '彩票', ]

pro_urls = []


"""全局变量定义"""
EXCEL_PATH = None  # excel文件路径
# RANGE_FIRST  # 循环起始行
# RANGE_LAST   # 循环结束行


# ----------------------------------------工作线程代码块--------------------------------------------------------------------------
class WorkerThread(QThread):
    update_signal = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.url_id = 0
        self.urls = []

    def run(self):
        self.update_signal.emit("----------------开始--------------")
        self.url_id, self.urls = read_excel()
        print(self.urls)
        tasks = []
        url_id = 1
        asyncio.set_event_loop(asyncio.new_event_loop())  # 创建新的事件循环
        loop = asyncio.get_event_loop()
        asyncio.set_event_loop(loop)
        for url in self.urls:
            # url = "http://" + url
            # 创建协程对象
            c = self.get_request(url_id, url)
            # 创建任务对象
            task = asyncio.ensure_future(c)
            # 绑定回调
            task.add_done_callback(self.parse)
            tasks.append(task)
            url_id = url_id + 1

        # loop = asyncio.get_event_loop()
        loop.run_until_complete(asyncio.wait(tasks))
        # 事件循环终止时的清
        # loop.close()
        self.update_signal.emit("----------------结束--------------")

    """异步获取URL信息"""
    async def get_request(seLf, url_id, url):
        connector = aiohttp.TCPConnector(ssl=False)
        async with aiohttp.ClientSession(connector=connector) as session:
            try:
                async with await session.get(url=url, headers=headers, timeout=10) as response:
                    try:
                        html = await response.text()
                        status_code = response.status
                        return url_id, url, status_code, html
                    except UnicodeDecodeError as ce:
                        try:
                            html = await response.text(encoding="gb2312")
                            status_code = response.status
                            return url_id, url, status_code, html
                        except UnicodeDecodeError as xe:
                            html = f"，解码发生错误：{xe}"
                            return url_id, url, None, html
                    # except Exception as e:
                    #     print(f"{url}，解码发生未知错误{e}")
                    #     return None
            except aiohttp.ClientError as ce:
                # 捕获客户端错误，如连接问题、DNS 解析问题、证书验证失败等
                html = f"客户端错误: {ce}"
                pro_urls.append(url)
                return url_id, url, None, html
            except asyncio.TimeoutError:
                # 处理超时错误
                html = f"请求超时"
                pro_urls.append(url)
                return url_id, url, None, html
            except Exception as e:
                pro_urls.append(url)
                html = f"未知异常: {e}"
                return url_id, url, None, html

    def parse(self, task):
        url_id, url, status_code, html = task.result()
        web_status = "正常网站"
        if status_code:
            """获取网站信息"""
            # 获取标题
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.title.string if soup.title else 'No title found'
            # 去除标题空格
            try:
                title = title.strip()
            except AttributeError:
                pass  # 对象没有strip属性

            """判断网站是否异常或违法"""
            if status_code == 200 and title:
                for item in Errors:
                    if title.find(item) != -1:  # 如果title中包含dir的某个元素
                        web_status = "异常网站"
                        break
                for item in Illegal:
                    if title.find(item) != -1:
                        web_status = "违法网站"
                        break

                # 提取域名的正则表达式
                domain_pattern = re.compile(r'https?://([^/]+)')
                # 使用正则表达式匹配域名
                match = domain_pattern.match(url)
                if match.group(1) in title and "官网首页" in title:
                    web_status = "域名出售"
            else:
                web_status = "异常网站"

            """获取备案信息"""
            # 获取icp备案
            icp_number = parse_icp(soup)
            # 获取公安备案
            public_security_number = parse_public_security(soup)

            ret_info = "-------------------------------\n"
            ret_info += f"网站情况：{web_status}\n"
            ret_info += f"URL：{url}\n"
            ret_info += f"标题：{title}\n"
            ret_info += f"状态码：{status_code}\n"
            ret_info += f"icp备案：{icp_number}\n"
            ret_info += f"公安备案：{public_security_number}\n"

            values = [url, web_status, title, status_code, icp_number, public_security_number]
            output_excel(url_id, values)
            self.update_signal.emit(ret_info)

        else:
            web_status = "错误网站"

            ret_info = "-------------------------------\n"
            ret_info += f"网站情况：{web_status}\n"
            ret_info += f"URL：{url}\n"
            self.update_signal.emit(ret_info)

            values = [url, web_status, html]
            output_excel(url_id, values)


# ----------------------------------------窗口代码块--------------------------------------------------------------------------
class MyWindow(QWidget, Ui_widget):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.bind()

    def bind(self):
        #self.btn.clicked.connect(self.func)

        # 创建工作线程
        self.worker_thread = WorkerThread()
        self.worker_thread.update_signal.connect(self.update_text_edit)

        # 开始运行按钮
        self.start_btn.clicked.connect(self.start_btn_func)
        # 结束运行按钮
        self.exit_btn.clicked.connect(QApplication.instance().quit)
        # 选择表格 读取urls
        self.excel_select_btn.clicked.connect(self.excel_select_btn_func)

    """启动工作线程"""
    def start_worker_thread(self):
        # 启动工作线程
        self.worker_thread.start()

    """"实时更新文本框显示内容"""
    def update_text_edit(self, message):
        # 在文本显示框中实时更新代码运行情况
        self.operation_status_edit.append(message)

    """选择表格按钮函数"""
    def excel_select_btn_func(self):
        # 打开文件对话框以获取Excel文件路径
        file_dialog = QFileDialog()
        file_path = file_dialog.getOpenFileName(self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls);;所有文件 (*)")
        self.excel_input_edit.setText(file_path[0])
        global EXCEL_PATH
        EXCEL_PATH = file_path[0]

    """开始按钮函数"""
    def start_btn_func(self):
        if not EXCEL_PATH:
            print("未选择文件路径")
        else:
            # main_start()
            self.start_worker_thread()


# ----------------------------------------爬虫代码块--------------------------------------------------------------------------

""""从Excel表中读取URL"""
def read_excel():
    url_id = 0
    urls = []
    try:
        # 打开 Excel 文件
        # print(view.EXCEL_PATH)
        workbook = openpyxl.load_workbook(EXCEL_PATH)
        # 选择第一个工作表
        sheet = workbook.active
        # 遍历行
        for row in sheet.iter_rows(values_only=True):
            url_id = url_id + 1
            url = "http://" + row[0]
            urls.append(url)
    except FileNotFoundError:
        print("文件未找到，请检查文件路径是否正确。")
    except Exception as e:
        print(f"发生了其他错误: {e}")
    return url_id, urls


"""icp备案获取"""
def parse_icp(soup):
    # 找到包含 ICP 备案关键字的元素
    icp_elements = soup.find_all(string=re.compile(r'ICP备\d'))
    # 提取备案信息
    icp_number = "未找到icp备案"
    for element in icp_elements:
        # 进一步处理字符串，提取具体备案信息
        # icp_match = re.search(r'ICP备[^\d]*(\d+)[^\u4e00-\u9fa5]*([\u4e00-\u9fa5]+)', element)
        # icp_match = re.search(r'([\u4e00-\u9fa5]*ICP备[^\d]*\d+[^\u4e00-\u9fa5]*[\u4e00-\u9fa5]+-*\d*)', element)
        icp_match = re.search(r'([\u4e00-\u9fa5]?ICP备\d+[^\u4e00-\u9fa5]*[\u4e00-\u9fa5]+-*\d*)', element)
        if icp_match:
            icp_number = icp_match.group(1)
        else:
            icp_number = "可能存在icp备案，但未爬取到"
    icp_number = icp_number.replace(" ", "")
    return icp_number


"""公安备案获取"""
def parse_public_security(soup):
    # 找到包含公安备案关键字的元素
    public_security_elements = soup.find_all(string=re.compile(r'公网安备'))
    # 提取公安备案信息
    public_security_number = "未找到公安备案"
    for element in public_security_elements:
        # 进一步处理字符串，提取具体备案信息
        public_security_match = re.search(r'([\u4e00-\u9fa5]?公网安备\s*\d+号)', element)
        if public_security_match:
            public_security_number = public_security_match.group(1)
        else:
            public_security_number = "可能存在公安备案，但未爬取到"
    public_security_number = public_security_number.replace(" ", "")
    return public_security_number


"""列表分割函数"""
def split_into_lists(lst, chunk_size):
    """将列表按照指定大小分割成子列表"""
    for i in range(0, len(lst), chunk_size):
        yield lst[i:i + chunk_size]


"""使用表格并写入"""
def output_excel(url_id, values):
    # 打开已存在的 Excel 文件
    try:
        workbook = openpyxl.load_workbook("../output.xlsx")
        # 获取要写入的工作表
        sheet = workbook.active  # 这里假设你要操作的是默认的工作表
        col_index = 0
        for value in values:
            row_index = url_id
            col_index = col_index + 1
            sheet.cell(row=row_index, column=col_index, value=value)
            print(value)
        # 保存工作簿
        workbook.save('output.xlsx')
    except FileNotFoundError:
        print("未找到文件，为您创建新文件存储。")
        create_excel(url_id, values)
    except Exception as e:
        print(f"发生了其他错误: {e}")


"""创建表格并写入"""
def create_excel(url_id, values):
    try:
        # 创建一个新的 Excel 工作簿
        workbook = openpyxl.Workbook()
        # 获取默认的工作表
        sheet = workbook.active
        col_index = 0
        for value in values:
            row_index = url_id
            col_index = col_index + 1
            sheet.cell(row=row_index, column=col_index, value=value)
            print(value)
        # 保存工作簿
        workbook.save('../output.xlsx')
    except FileNotFoundError:
        print("文件未找到，请检查文件路径是否正确。")
    except Exception as e:
        print(f"发生了其他错误: {e}")







if __name__ == '__main__':
    app = QApplication([])
    window = MyWindow()
    window.show()
    app.exec()

